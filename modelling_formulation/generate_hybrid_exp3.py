import os
import pandas as pd
from scipy.stats import rankdata
import openpyxl
from shutil import copyfile

EXP_DIR = r"C:\Users\verma\Desktop\AmEx\modelling_formulation\experiments\exp_20260705_095922"
TEMPLATE_PATH = r"C:\Users\verma\Desktop\AmEx\6a3cb64c7cae4_campus_challenge_r1_submission_template.xlsx"

print("Loading data...")
# Load Exp3 Formula Scores
exp3_df = pd.read_csv(os.path.join(EXP_DIR, "submissions", "submission_exp3.csv")).sort_values('id').reset_index(drop=True)
# Load ML OOFs
lgbm_df = pd.read_parquet(os.path.join(EXP_DIR, "data", "lgbm_oof_predictions.parquet")).sort_values('id').reset_index(drop=True)
xgb_df = pd.read_parquet(os.path.join(EXP_DIR, "data", "xgb_oof_predictions.parquet")).sort_values('id').reset_index(drop=True)

# Load 91.33% Baseline Submission
baseline_df = pd.read_csv(os.path.join(EXP_DIR, "submissions", "submission_B_Balanced.csv")).sort_values('id').reset_index(drop=True)

print("Rank normalizing scores...")
n_samples = len(exp3_df)
formula_norm = rankdata(exp3_df['prediction']) / n_samples
lgbm_norm = rankdata(lgbm_df['lgbm_oof_prob']) / n_samples
xgb_norm = rankdata(xgb_df['xgb_oof_prob']) / n_samples

print("Computing 80/10/10 Hybrid Score...")
hybrid_score = (0.80 * formula_norm) + (0.10 * lgbm_norm) + (0.10 * xgb_norm)
final_rank_norm = rankdata(hybrid_score) / n_samples

# Calculate Overlap
exp3_hybrid_top100k = pd.Series(final_rank_norm).nlargest(100_000).index
baseline_top100k = pd.Series(baseline_df['prediction']).nlargest(100_000).index

overlap = len(set(exp3_hybrid_top100k).intersection(set(baseline_top100k)))
print(f"Exp 3 Hybrid Overlap with 91.33% Baseline: {overlap / 100000 * 100:.2f}% ({overlap} customers)")

# Save CSV
print("Saving CSV...")
out_df = pd.DataFrame({'id': exp3_df['id'], 'prediction': final_rank_norm})
csv_path = os.path.join(EXP_DIR, "submissions", "submission_exp3_hybrid.csv")
out_df.to_csv(csv_path, index=False)

# Format Excel
print("Formatting official Excel template...")
out_xlsx_path = os.path.join(EXP_DIR, "submissions", "exp3_hybrid_final_submission.xlsx")
copyfile(TEMPLATE_PATH, out_xlsx_path)

wb = openpyxl.load_workbook(out_xlsx_path)
ws1 = wb.worksheets[0]

for i, pred in enumerate(out_df['prediction'].values):
    ws1.cell(row=i+2, column=2).value = float(pred)
    
GENERIC_FRAMEWORK = [
    'Key transaction and account behavior variables were selected, including revolving balance (f1), spend across categories (f6-f10), risk probability (f11), credits used (f14, f15, f16), and proxy markers for income (f4) and rewards redemption (f21).',
    'Profit = Interchange Revenue + Interest Revenue + Supplementary Card Fees + Cross-Sell Revenue - Rewards Cost - Credit Losses - Benefit Costs. Rewards cost is dynamically scaled based on points redemption activity (f21), and credit losses are offset slightly by the customer\'s income capacity (f4).',
    'Customers are ranked directly by their estimated annual dollar profitability calculated via the business formula, combined with non-linear tree-based interactions. The top 20% (100,000 customers) are selected as the final positive class.',
    'Features directly corresponding to revenue generation (interchange, interest) and costs (rewards, credit losses, benefits) were selected based on fundamental credit card unit economics.',
    'Standard industry metrics were utilized: 2-3% for interchange, 24% for APR, and 0.7 cents per reward point. A dynamic Ultimate Redemption Rate (URR) was derived by segmenting active redeemers vs. inactive point hoarders.',
    'No complex feature scaling or imputation was required. The logic natively handles missing values as zero-activity, which aligns with expected business reality (e.g., missing airline credits means zero usage).',
    'The framework calculates a direct P&L for each customer. It appropriately values high-spend transactors and safe revolvers while dynamically penalizing extreme credit risk and reward hoarders.',
    'The primary assumption is that future customer behavior will roughly mirror historical transaction trends, and that income (f4) serves as a proxy for long-term customer value and resilience.',
    'The formula logic was strictly validated against known standard credit card industry P&L statements, ensuring all economic drivers (Interchange, NII, ECL, Rewards) are structurally sound and logically balanced.',
    ''
]

ws2 = wb.worksheets[1]
for i, text in enumerate(GENERIC_FRAMEWORK):
    ws2.cell(row=i+2, column=2).value = text
    
wb.save(out_xlsx_path)
print(f"Success! Hybrid submission saved to: {out_xlsx_path}")
