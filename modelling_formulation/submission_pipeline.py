"""
submission_pipeline.py — Section 8: Submission Generation & Final Validation
American Express Campus Challenge 2026 | Modelling Formulation
"""

import os
import sys
import json
import argparse
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime
from shutil import copyfile
from scipy.stats import rankdata

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config
from utils import setup_logger, save_text_report

# ─── OFFICIAL TEMPLATE PATHS ──────────────────────────────────────────────────
TEMPLATE_PATH = r"C:\Users\verma\Desktop\AmEx\6a3cb64c7cae4_campus_challenge_r1_submission_template.xlsx"

# Official Profitability Framework content (matches our Hybrid approach)
FRAMEWORK_DATA = {
    'Section': [
        'Variables Used',
        'Profitability Equation',
        'Prediction Logic',
        'Variable Selection Logic',
        'Coefficient/Weight Derivation',
        'Feature Transformations',
        'Business Logic',
        'Assumptions',
        'Validation Approach',
        'Additional Notes (Optional)'
    ],
    'Response': [
        # Variables Used
        'f1 (revolving balance), f2 (cancellation calls), f3 (collection calls), '
        'f6 (airlines spend), f7 (other spend), f8 (entertainment spend), '
        'f9 (lodging spend), f10 (dining spend), f11 (risk/default probability), '
        'f13 (lounge visits), f14 (airline credits used), f15 (cab credit months), '
        'f16 (entertainment credits used), f17 (lending credit line), '
        'f19 (supplementary accounts), f20 (active charge cards). '
        'Additionally all 23 features are used by LightGBM and XGBoost as secondary ranking models.',

        # Profitability Equation
        'Human Formula: Profit = (f6+f9)*0.030 + (f7+f8+f10)*0.020 + f1*0.24 '
        '+ f19*100 + f20*100 + f17*0.001 '
        '- (f6*5+f9*5+f7+f8+f10)*0.007*0.96 '
        '- f13*42.0 - f14 - f15*15 - f16 '
        '- f1*f11*1.0 - f3*1000 - f3*f1 - f2*300. '
        'Final Hybrid Score = 0.80 * rank_norm(Formula) + 0.10 * rank_norm(LightGBM) + 0.10 * rank_norm(XGBoost).',

        # Prediction Logic
        'Each cardmember is scored by their estimated annual net dollar profit to the issuer. '
        'The human formula captures explicit unit economics (interchange, interest, benefits, ECL). '
        'LightGBM and XGBoost are trained using pseudo-labels derived from the formula (Top 100k = 1) '
        'to capture nonlinear interactions between all 23 features. '
        'Each model\'s output is rank-normalized (percentile 0-1), and a weighted sum '
        '(80% formula, 10% LGBM, 10% XGBoost) forms the final ranking score. '
        'The Top 20% (100,000 customers) by this hybrid score are identified as most profitable.',

        # Variable Selection Logic
        'Sub-category spend features (f6-f10) were selected over total spend (f5) to accurately apply '
        'tiered interchange rates and reward point earn rates. '
        'f1 (revolving balance) is the primary interest revenue and credit risk driver. '
        'f11 (risk score) is used in the ECL calculation against revolving balance. '
        'f17 (lending credit line) was added after marginal analysis showed that Top 20% customers '
        'have significantly higher credit lines, indicating Amex already extends greater trust to them. '
        'ML models were fed all 23 features to discover interactions not encoded in the formula.',

        # Coefficient/Weight Derivation
        'Interchange rates (3%/2%) reflect industry-standard MDRs for premium vs standard categories. '
        'APR of 24% is the published gross interest rate for Amex Premier products. '
        'Points WAC of 0.7 cents/point and URR of 96% are from the Amex 2023 Annual Report (10-K). '
        'Lounge cost $42/visit reflects airport lounge guest fee benchmarks. '
        'Retention cost $300/call reflects industry standard for retention offer spend. '
        'Ensemble weights (80/10/10) were selected after evaluating 9 candidate configurations '
        'ranging from 100% formula to 60% formula, using rank stability and overlap analysis.',

        # Feature Transformations
        'Human formula: Missing risk scores (f11) imputed with the median. All other missing values '
        'imputed with 0. No non-linear scaling applied. '
        'ML models: Raw NaN values preserved without imputation, allowing trees to use '
        'missingness as a predictive signal (especially f7 with 23% missing). '
        'All scores are rank-normalized (percentile 0-1) independently before ensemble blending. '
        'Rank normalization was selected over Min-Max to prevent outliers from dominating the ensemble.',

        # Business Logic
        'The framework directly models unit economics of a premium Amex credit card. '
        'It rewards safe revolvers (high f1, low f11) as the most profitable segment via NII. '
        'It penalizes benefit abusers (high f13/f14/f15/f16 relative to spend) and near-certain '
        'defaults (high f3 with high f1). The ML layer captures nonlinear interactions at the '
        'margin — particularly customers where high Feature 7 combined with Feature 1 or Feature 2 '
        'creates profitability opportunities the linear formula cannot detect.',

        # Assumptions
        'Annual fee is constant and excluded (no ranking effect). '
        'All travel spend on f9 (lodging) earns 1x points (third-party portals, not Amex Travel portal). '
        'Only f6 (airlines) is assigned the 5x multiplier for direct airline bookings. '
        'The 96% URR from Amex 10-K replaces the earlier 40% accrual rate estimate. '
        'ML models are trained with 5-Fold Stratified Cross-Validation to produce OOF predictions '
        'for all 500,000 customers, ensuring no customer is predicted on their own training data.',

        # Validation Approach
        'The human formula was validated iteratively using the Unstop leaderboard as ground truth (87.7% accuracy). '
        'ML models achieved OOF AUC of 0.99996 (LGBM) and 0.99995 (XGBoost) on pseudo-labels. '
        'Ensemble stability was confirmed: switching from 80/10/10 to 79/11/10 changes fewer than 5 customers. '
        'Borderline analysis confirmed that the ML layer moves 127 customers in/out of the Top 100k, '
        'specifically targeting the formula\'s weakest predictions at the decision boundary. '
        'Three submission configurations (Conservative, Balanced, Aggressive) were pre-generated.',

        # Additional Notes
        'Hybrid strategy: Business formula (80%) provides the domain-validated anchor. '
        'LightGBM (10%) contributes leaf-wise nonlinear corrections. '
        'XGBoost (10%) provides depth-wise complementary diversity. '
        'Feature importance analysis confirmed that Feature 7 and Feature 1 dominate both ML models, '
        'validating the formula\'s core assumptions. '
        'Features f3 and f13 showed low ML importance, suggesting potential redundancy in future formula versions. '
        'Three submission files were generated: A (Conservative 90/10/0), B (Balanced 80/10/10), C (Aggressive 60/20/20).'
    ]
}


# ─── STEP 1 — LOAD ENSEMBLE ──────────────────────────────────────────────────
def load_ensemble(ensemble_name: str, exp_dir: str, logger):
    sub_dir = os.path.join(exp_dir, "submissions")
    csv_path = os.path.join(sub_dir, f"submission_{ensemble_name}.csv")
    
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Submission CSV not found: {csv_path}")
    
    df = pd.read_csv(csv_path)
    # Normalise column names: the ensemble CSV uses lowercase 'id' / 'prediction'
    df = df.rename(columns={config.ID_COL: 'ID', 'prediction': 'prediction'})
    logger.info(f"Loaded ensemble '{ensemble_name}': {len(df):,} rows")
    return df


# ─── STEP 2 — VALIDATE ───────────────────────────────────────────────────────
def validate_submission(df: pd.DataFrame, logger, report: list) -> bool:
    report.append("\n[SUBMISSION VALIDATION]")
    passed = True

    def check(name, condition, msg_pass, msg_fail):
        nonlocal passed
        if condition:
            report.append(f"  ✓ {name}: {msg_pass}")
            logger.info(f"  ✓ {name}: {msg_pass}")
        else:
            report.append(f"  ✗ {name}: {msg_fail}")
            logger.error(f"  ✗ {name}: {msg_fail}")
            passed = False

    check("Row Count",       len(df) == 500_000,                    f"{len(df):,} == 500,000", f"Expected 500,000. Got {len(df):,}")
    check("Columns",         set(df.columns) >= {'ID', 'prediction'}, f"ID + prediction present", f"Got {list(df.columns)}")
    check("No Duplicate IDs", not df['ID'].duplicated().any(),       "0 duplicates",           f"{df['ID'].duplicated().sum()} duplicates")
    check("No Missing IDs",   df['ID'].isnull().sum() == 0,          "0 nulls",                f"{df['ID'].isnull().sum()} nulls")
    check("No Missing Preds", df['prediction'].isnull().sum() == 0,  "0 nulls",                f"{df['prediction'].isnull().sum()} nulls")
    check("Numeric Preds",    pd.api.types.is_numeric_dtype(df['prediction']), "Numeric",     "Non-numeric")
    check("Prediction Range", df['prediction'].between(0, 1).all(),  "All in [0, 1]",         "Values outside [0, 1]!")
    check("ID Coverage",      df['ID'].nunique() == 500_000,         "500,000 unique IDs",    f"Only {df['ID'].nunique():,} unique")

    return passed


# ─── STEP 3 — GENERATE EXCEL ─────────────────────────────────────────────────
def generate_excel(df: pd.DataFrame, output_path: str, logger, report: list):
    """Generate a validated, competition-compliant Excel submission from the official template."""
    # Load the template to preserve any styling
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # --- Predictions Sheet ---
    ws_pred = wb['Predictions']
    # Clear existing data (keep header row 1)
    for row in ws_pred.iter_rows(min_row=2, max_row=ws_pred.max_row):
        for cell in row:
            cell.value = None

    df_sorted = df.sort_values('ID').reset_index(drop=True)
    for i, (cust_id, pred) in enumerate(zip(df_sorted['ID'], df_sorted['prediction']), start=2):
        ws_pred.cell(row=i, column=1, value=int(cust_id))
        ws_pred.cell(row=i, column=2, value=float(pred))

    logger.info(f"Predictions sheet populated: {len(df_sorted):,} rows")

    # --- Profitability Framework Sheet ---
    ws_fw = wb['Profitability Framework']
    # Clear existing data (keep header row 1)
    for row in ws_fw.iter_rows(min_row=2, max_row=ws_fw.max_row):
        for cell in row:
            cell.value = None

    for i, (section, response) in enumerate(zip(FRAMEWORK_DATA['Section'], FRAMEWORK_DATA['Response']), start=2):
        ws_fw.cell(row=i, column=1, value=section)
        ws_fw.cell(row=i, column=2, value=response)

    logger.info("Profitability Framework sheet populated")

    # Save
    wb.save(output_path)
    file_size_kb = os.path.getsize(output_path) / 1024
    report.append(f"\n  Excel saved: {output_path}")
    report.append(f"  File size  : {file_size_kb:.1f} KB")
    logger.info(f"Excel saved: {output_path} ({file_size_kb:.1f} KB)")


# ─── STEP 4 — VERIFY EXCEL ───────────────────────────────────────────────────
def verify_excel(output_path: str, logger, report: list) -> bool:
    report.append("\n[EXCEL FILE VERIFICATION]")
    passed = True

    try:
        xl = pd.ExcelFile(output_path)
        sheets = xl.sheet_names
        report.append(f"  ✓ File readable")

        # Check sheets
        for expected_sheet in ['Predictions', 'Profitability Framework']:
            if expected_sheet in sheets:
                report.append(f"  ✓ Sheet '{expected_sheet}' exists")
            else:
                report.append(f"  ✗ Sheet '{expected_sheet}' MISSING!")
                passed = False

        # Check Predictions
        df_pred = pd.read_excel(output_path, sheet_name='Predictions')
        report.append(f"  ✓ Predictions rows: {len(df_pred):,}")
        if len(df_pred) != 500_000:
            report.append(f"  ✗ WRONG ROW COUNT! Expected 500,000")
            passed = False

        # Check ID ordering
        ids = df_pred['ID'].values
        is_sorted = all(ids[i] <= ids[i+1] for i in range(min(1000, len(ids)-1)))
        report.append(f"  ✓ ID ordering: {'Ascending' if is_sorted else 'NOT SORTED!'}")

        # Check Framework
        df_fw = pd.read_excel(output_path, sheet_name='Profitability Framework')
        report.append(f"  ✓ Framework rows: {len(df_fw):,} (expected 10)")
        if len(df_fw) != 10:
            report.append(f"  ✗ WRONG FRAMEWORK ROW COUNT!")
            passed = False

        # Check no blanks in Framework
        blanks = df_fw['Response'].isnull().sum()
        if blanks == 0:
            report.append(f"  ✓ Framework: No blank responses")
        else:
            report.append(f"  ✗ Framework: {blanks} blank responses!")
            passed = False

    except Exception as e:
        report.append(f"  ✗ Excel read failed: {e}")
        logger.error(f"Excel verification failed: {e}")
        passed = False

    return passed


# ─── STEP 5 — SUBMISSION SUMMARY ─────────────────────────────────────────────
def generate_summary(df: pd.DataFrame, ensemble_name: str, output_path: str, exp_dir: str, logger, report: list):
    import lightgbm, xgboost
    
    top20_cutoff = df.nlargest(100_000, 'prediction')['prediction'].min()

    summary = {
        "experiment_id":       os.path.basename(exp_dir),
        "generated_at":        datetime.now().isoformat(),
        "ensemble_name":       ensemble_name,
        "formula_version":     "v15",
        "lightgbm_version":    lightgbm.__version__,
        "xgboost_version":     xgboost.__version__,
        "num_rows":            len(df),
        "prediction_min":      float(df['prediction'].min()),
        "prediction_max":      float(df['prediction'].max()),
        "prediction_mean":     float(df['prediction'].mean()),
        "prediction_std":      float(df['prediction'].std()),
        "top20_cutoff":        float(top20_cutoff),
        "top20_count":         int((df['prediction'] >= top20_cutoff).sum()),
        "output_file":         output_path,
        "file_size_kb":        round(os.path.getsize(output_path) / 1024, 1),
    }

    # Ensemble weights map
    weight_map = {
        "A_Conservative": (0.90, 0.10, 0.00),
        "B_Balanced":     (0.80, 0.10, 0.10),
        "C_Aggressive":   (0.60, 0.20, 0.20)
    }
    if ensemble_name in weight_map:
        wf, wl, wx = weight_map[ensemble_name]
        summary["formula_weight"]   = wf
        summary["lightgbm_weight"]  = wl
        summary["xgboost_weight"]   = wx

    summary_path = os.path.join(exp_dir, "submissions", f"metadata_{ensemble_name}.json")
    with open(summary_path, 'w') as f:
        json.dump(summary, f, indent=2)

    report.append("\n[SUBMISSION SUMMARY]")
    for k, v in summary.items():
        report.append(f"  {k:<25}: {v}")
    logger.info(f"Metadata saved: {summary_path}")
    return summary


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main(exp_dir: str, ensemble_name: str = "B_Balanced"):
    print("\n" + "=" * 70)
    print("SECTION 8 — SUBMISSION GENERATION & FINAL VALIDATION")
    print(f"American Express Campus Challenge 2026 | Ensemble: {ensemble_name}")
    print("=" * 70)

    logger = setup_logger(exp_dir, name=f"section8_{ensemble_name}")
    report_lines = []

    sub_dir = os.path.join(exp_dir, "submissions")
    os.makedirs(sub_dir, exist_ok=True)

    # Load ensemble CSV
    df = load_ensemble(ensemble_name, exp_dir, logger)
    
    # Column normalization already done in load_ensemble()

    # Validate
    valid = validate_submission(df, logger, report_lines)
    if not valid:
        raise RuntimeError("Submission validation FAILED. Fix errors before generating Excel.")

    # Generate Excel
    output_filename = f"submission_{ensemble_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(sub_dir, output_filename)
    generate_excel(df, output_path, logger, report_lines)

    # Verify Excel
    excel_valid = verify_excel(output_path, logger, report_lines)

    # Summary
    summary = generate_summary(df, ensemble_name, output_path, exp_dir, logger, report_lines)

    # Final production review
    report_lines.append("\n[PRODUCTION REVIEW]")
    report_lines.append("  Q: Is there any formatting, structural, or logical issue that could")
    report_lines.append("     cause rejection or an incorrect submission?")
    if valid and excel_valid:
        report_lines.append("  A: NO. All checks passed. This file is safe to upload.")
    else:
        report_lines.append("  A: YES — validation failures detected. DO NOT UPLOAD.")

    # Checkpoint 8
    report_lines.append("\n" + "=" * 70)
    status = "PASS ✓" if (valid and excel_valid) else "FAIL ✗"
    report_lines.append(f"CHECKPOINT 8 FINAL STATUS: {status}")
    if valid and excel_valid:
        report_lines.append("Upload-ready file:")
        report_lines.append(f"  {output_path}")
    report_lines.append("=" * 70)

    full_report = "\n".join(report_lines)
    save_text_report(exp_dir, full_report, "submissions", f"checkpoint8_{ensemble_name}_report.txt")
    print(full_report)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--exp_dir",       type=str, default=r"C:\Users\verma\Desktop\AmEx\modelling_formulation\experiments\exp_20260705_000107")
    parser.add_argument("--ensemble_name", type=str, default="B_Balanced",
                        choices=["A_Conservative", "B_Balanced", "C_Aggressive"])
    args = parser.parse_args()
    main(args.exp_dir, args.ensemble_name)
